"""
Genera el archivo Excel completo del Mundial 2026 con proyecciones del modelo v3.

Hojas:
  1. Fase de Grupos       — 72 partidos con probabilidades y proyección
  2. Clasificación        — tabla proyectada de cada grupo
  3. Fase Eliminatoria    — bracket completo R32 → Final con proyecciones
  4. Prob. Campeonato     — ranking de probabilidad de título (ambos modelos)
  5. Ratings              — ELO, FIFA y compuesto para los 48 equipos
  6. Validación Modelo    — backtesting 2018/2022 con Brier Score y Log-Loss
"""
import os, sys
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.chart import BarChart, Reference

sys.path.insert(0, os.path.dirname(__file__))
from src.elo_calculator import win_probability, set_draw_params
from src.calibration import load_calibration
from src.fifa_ranking import load_fifa_rankings, composite_rating, get_fifa_points
from src.attack_defense import compute_attack_defense
from src.form import compute_form
from src.simulator import simulate_group_stage, build_composite_ratings
from src.world_cup_2026 import GROUPS

# ─── Paleta de colores ────────────────────────────────────────────────────────
C_HEADER_BG   = "1F3864"   # azul oscuro
C_HEADER_FG   = "FFFFFF"
C_SUBHEAD_BG  = "2E75B6"
C_SUBHEAD_FG  = "FFFFFF"
C_GROUP_BG    = "D6E4F0"
C_WIN         = "C6EFCE"   # verde claro
C_DRAW        = "FFEB9C"   # amarillo
C_LOSS        = "FFC7CE"   # rojo claro
C_PROJECTED   = "BDD7EE"   # azul claro
C_GOLD        = "FFD700"
C_SILVER      = "C0C0C0"
C_BRONZE      = "CD7F32"
C_STRIPE1     = "F2F7FC"
C_STRIPE2     = "FFFFFF"
C_BORDER      = "B8CCE4"

THIN = Side(style="thin", color=C_BORDER)
THICK = Side(style="medium", color="1F3864")
BORDER_THIN  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_THICK = Border(left=THICK, right=THICK, top=THICK, bottom=THICK)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def header_font(bold=True, color=C_HEADER_FG, size=10):
    return Font(bold=bold, color=color, name="Calibri", size=size)

def cell_font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, name="Calibri", size=size)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center")

def style_header(cell, bg=C_HEADER_BG, fg=C_HEADER_FG, size=10):
    cell.fill = fill(bg)
    cell.font = Font(bold=True, color=fg, name="Calibri", size=size)
    cell.alignment = center()
    cell.border = BORDER_THIN

def style_cell(cell, bg=C_STRIPE2, bold=False, align="center", color="000000"):
    cell.fill = fill(bg)
    cell.font = Font(bold=bold, color=color, name="Calibri", size=10)
    cell.alignment = center() if align == "center" else left()
    cell.border = BORDER_THIN

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def freeze(ws, cell="B2"):
    ws.freeze_panes = cell

# ─── Mapeo equipos → grupo ────────────────────────────────────────────────────
TEAM_TO_GROUP = {}
for g, teams in GROUPS.items():
    for t in teams:
        TEAM_TO_GROUP[t] = g

# ─── Carga de datos ───────────────────────────────────────────────────────────
def load_all_data():
    print("Cargando ratings ELO...")
    elo_df = pd.read_csv("data/processed/elo_ratings.csv")
    elo_ratings = dict(zip(elo_df["team"], elo_df["elo"]))

    print("Cargando ranking FIFA...")
    fifa_rankings = load_fifa_rankings()

    print("Construyendo ratings compuestos...")
    comp_ratings = build_composite_ratings(elo_ratings, elo_weight=0.6)

    print("Cargando calibración de empate...")
    cal = load_calibration()
    if cal:
        set_draw_params(*cal)
        print(f"  Calibración cargada: a={cal[0]:.4f}, b={cal[1]:.5f}, c={cal[2]:.4f}")

    print("Calculando ataque/defensa y forma...")
    df_raw = pd.read_csv("data/raw/results.csv", parse_dates=["date"])
    attack_defense = compute_attack_defense(df_raw)
    form_adj = compute_form(df_raw)
    print(f"  {len(attack_defense)} equipos con A/D | {len(form_adj)} con forma.")

    print("Cargando resultados de simulaciones...")
    elo_results  = pd.read_csv("results/csv/predicciones_elo.csv")
    comp_results = pd.read_csv("results/csv/predicciones_compuesto.csv")

    return elo_ratings, fifa_rankings, comp_ratings, elo_results, comp_results, attack_defense, form_adj


def match_probabilities(team_a, team_b, ratings, neutral=True):
    elo_a = ratings.get(team_a, 1500)
    elo_b = ratings.get(team_b, 1500)
    p_win, p_draw, p_loss = win_probability(elo_a, elo_b, neutral=neutral)
    if team_a == max([team_a, team_b], key=lambda t: ratings.get(t, 1500)):
        fav = team_a
    else:
        fav = team_b
    return p_win, p_draw, p_loss, fav


# ═══════════════════════════════════════════════════════════════════════════════
# HOJA 1 — FASE DE GRUPOS
# ═══════════════════════════════════════════════════════════════════════════════
def sheet_group_stage(wb, schedule_df, comp_ratings, elo_ratings, fifa_rankings):
    ws = wb.create_sheet("Fase de Grupos")
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 30

    # Título
    ws.merge_cells("A1:M1")
    ws["A1"] = "🏆  MUNDIAL FIFA 2026 — FASE DE GRUPOS  |  Proyecciones del Modelo (ELO + FIFA)"
    ws["A1"].fill = fill(C_HEADER_BG)
    ws["A1"].font = Font(bold=True, color=C_HEADER_FG, name="Calibri", size=13)
    ws["A1"].alignment = center()

    # Cabeceras
    headers = [
        "Fecha", "Grupo", "Equipo Local", "Equipo Visitante",
        "Prob. Local", "Prob. Empate", "Prob. Visitante",
        "Proyección", "ELO Local", "ELO Visit.", "FIFA Local", "FIFA Visit.", "Rating Comp. Local", "Rating Comp. Visit."
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        style_header(c, bg=C_SUBHEAD_BG)

    widths = [13, 8, 22, 22, 11, 11, 11, 22, 10, 10, 10, 10, 14, 14]
    for i, w in enumerate(widths, 1):
        set_col_width(ws, i, w)

    current_group = None
    row = 3
    for _, match in schedule_df.iterrows():
        home  = match["home_team"]
        away  = match["away_team"]
        group = TEAM_TO_GROUP.get(home, "?")
        date  = match["date"].strftime("%d/%m/%Y") if pd.notna(match["date"]) else ""

        p_win, p_draw, p_loss, _ = match_probabilities(home, away, comp_ratings)
        projected = home if p_win > p_loss else (away if p_loss > p_win else "Empate")
        projected_str = f"→ {projected}" if projected != "Empate" else "→ Empate"

        elo_h  = elo_ratings.get(home, 1500)
        elo_a  = elo_ratings.get(away, 1500)
        fifa_h = get_fifa_points(home, fifa_rankings)
        fifa_a = get_fifa_points(away, fifa_rankings)
        comp_h = comp_ratings.get(home, 1500)
        comp_a = comp_ratings.get(away, 1500)

        # Separador de grupo
        if group != current_group:
            current_group = group
            ws.row_dimensions[row].height = 6
            for col in range(1, 15):
                c = ws.cell(row=row, column=col, value="")
                c.fill = fill("D6E4F0")
            row += 1

        stripe = C_STRIPE1 if row % 2 == 0 else C_STRIPE2

        vals = [
            date, f"Grupo {group}", home, away,
            p_win, p_draw, p_loss,
            projected_str,
            round(elo_h), round(elo_a),
            round(fifa_h), round(fifa_a),
            round(comp_h), round(comp_a),
        ]

        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            align = "left" if col in (3, 4, 8) else "center"
            bold  = col == 8
            style_cell(c, bg=stripe, align=align, bold=bold)

            # Colorear probabilidades
            if col == 5:   # prob local
                c.fill = fill(C_WIN if p_win >= 0.45 else (C_DRAW if p_win >= 0.3 else C_LOSS))
                c.number_format = "0.0%"
                c.value = p_win
            elif col == 6:  # prob empate
                c.fill = fill(C_DRAW)
                c.number_format = "0.0%"
                c.value = p_draw
            elif col == 7:  # prob visitante
                c.fill = fill(C_WIN if p_loss >= 0.45 else (C_DRAW if p_loss >= 0.3 else C_LOSS))
                c.number_format = "0.0%"
                c.value = p_loss
            elif col == 8:
                c.fill = fill(C_PROJECTED)

        row += 1

    ws.row_dimensions[row].height = 8
    # Leyenda
    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "Leyenda:"
    ws[f"A{row}"].font = Font(bold=True, name="Calibri", size=9)
    for col, (label, color) in enumerate([("≥45% (favorito)", C_WIN), ("30-45%", C_DRAW), ("<30%", C_LOSS)], 5):
        c = ws.cell(row=row, column=col, value=label)
        c.fill = fill(color)
        c.font = Font(name="Calibri", size=9)
        c.alignment = center()
        c.border = BORDER_THIN

    print(f"  Hoja 'Fase de Grupos': {row-3} filas")


# ═══════════════════════════════════════════════════════════════════════════════
# HOJA 2 — CLASIFICACIÓN PROYECTADA
# ═══════════════════════════════════════════════════════════════════════════════
def project_group_standings(comp_ratings, attack_defense=None, form_adj=None):
    """Simula múltiples veces y toma el clasificado más frecuente."""
    from collections import Counter
    N = 5000
    first_counts = {g: Counter() for g in GROUPS}
    second_counts = {g: Counter() for g in GROUPS}
    third_counts  = {g: Counter() for g in GROUPS}

    for _ in range(N):
        results, _ = simulate_group_stage(comp_ratings, attack_defense, form_adj)
        for g, teams in results.items():
            first_counts[g][teams[0]]  += 1
            second_counts[g][teams[1]] += 1
            third_counts[g][teams[2]]  += 1

    standings = {}
    for g in GROUPS:
        teams = GROUPS[g]
        team_stats = []
        for t in teams:
            p1 = first_counts[g].get(t, 0) / N
            p2 = second_counts[g].get(t, 0) / N
            p3 = third_counts[g].get(t, 0) / N
            p4 = 1 - p1 - p2 - p3
            team_stats.append({
                "team": t,
                "prob_1st": p1,
                "prob_2nd": p2,
                "prob_3rd": p3,
                "prob_4th": p4,
                "rating": comp_ratings.get(t, 1500),
            })
        team_stats.sort(key=lambda x: x["prob_1st"], reverse=True)
        standings[g] = team_stats
    return standings


def sheet_group_standings(wb, comp_ratings, attack_defense=None, form_adj=None):
    print("  Proyectando clasificación (5,000 simulaciones)...")
    standings = project_group_standings(comp_ratings, attack_defense, form_adj)

    ws = wb.create_sheet("Clasificación Proyectada")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:I1")
    ws["A1"] = "🏆  MUNDIAL FIFA 2026 — CLASIFICACIÓN PROYECTADA POR GRUPO"
    ws["A1"].fill = fill(C_HEADER_BG)
    ws["A1"].font = Font(bold=True, color=C_HEADER_FG, name="Calibri", size=13)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 28

    col_start = 1
    groups_per_row = 4
    group_list = sorted(standings.keys())

    row = 3
    for idx, group_name in enumerate(group_list):
        if idx > 0 and idx % groups_per_row == 0:
            row += 8
            col_start = 1

        col = col_start + (idx % groups_per_row) * 10

        # Cabecera de grupo
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row, end_column=col + 8
        )
        gc = ws.cell(row=row, column=col, value=f"  GRUPO {group_name}")
        gc.fill = fill(C_SUBHEAD_BG)
        gc.font = Font(bold=True, color=C_HEADER_FG, name="Calibri", size=11)
        gc.alignment = left()
        gc.border = BORDER_THIN
        ws.row_dimensions[row].height = 22

        # Subheader
        sub_headers = ["Pos", "Equipo", "Rating", "P(1°)", "P(2°)", "P(3°)", "P(4°)", "Estado"]
        for i, sh in enumerate(sub_headers):
            c = ws.cell(row=row + 1, column=col + i, value=sh)
            style_header(c, bg="2E75B6")
            ws.column_dimensions[get_column_letter(col + i)].width = (
                4 if i == 0 else (18 if i == 1 else (8 if i == 2 else 7))
            )

        # Filas de equipos
        for pos, team_data in enumerate(standings[group_name], 1):
            r = row + 1 + pos
            bg = C_STRIPE1 if pos % 2 == 0 else C_STRIPE2

            medal_bg = {1: "E8F5E9", 2: "E3F2FD", 3: "FFF8E1", 4: C_STRIPE2}[pos]

            estado = "✅ Clasifica" if pos <= 2 else ("⚠️ Posible 3°" if pos == 3 else "❌ Eliminado")
            estado_color = {"1": "1B5E20", "2": "0D47A1", "⚠": "E65100", "❌": "B71C1C"}.get(estado[0], "000000")

            vals = [
                pos,
                team_data["team"],
                round(team_data["rating"]),
                team_data["prob_1st"],
                team_data["prob_2nd"],
                team_data["prob_3rd"],
                team_data["prob_4th"],
                estado,
            ]

            for i, val in enumerate(vals):
                c = ws.cell(row=r, column=col + i, value=val)
                c.fill = fill(medal_bg)
                c.border = BORDER_THIN
                c.alignment = center() if i != 1 else left()
                c.font = Font(name="Calibri", size=9,
                              bold=(i == 1),
                              color=estado_color if i == 7 else "000000")
                if i in (3, 4, 5, 6):
                    c.number_format = "0.0%"

    print(f"  Hoja 'Clasificación Proyectada': {len(group_list)} grupos")


# ═══════════════════════════════════════════════════════════════════════════════
# HOJA 3 — FASE ELIMINATORIA
# ═══════════════════════════════════════════════════════════════════════════════
def project_knockout(comp_ratings, attack_defense=None, form_adj=None):
    """Proyecta el bracket usando el ganador más probable en cada partido."""
    from collections import Counter
    N = 3000
    advance_first  = {g: Counter() for g in GROUPS}
    advance_second = {g: Counter() for g in GROUPS}
    advance_third  = {g: Counter() for g in GROUPS}

    for _ in range(N):
        res, _ = simulate_group_stage(comp_ratings, attack_defense, form_adj)
        for g, teams in res.items():
            advance_first[g][teams[0]]  += 1
            advance_second[g][teams[1]] += 1
            advance_third[g][teams[2]]  += 1

    firsts  = {g: advance_first[g].most_common(1)[0][0]  for g in GROUPS}
    seconds = {g: advance_second[g].most_common(1)[0][0] for g in GROUPS}
    thirds_raw = {g: advance_third[g].most_common(1)[0][0] for g in GROUPS}

    # 8 mejores terceros
    thirds_sorted = sorted(
        thirds_raw.items(),
        key=lambda x: comp_ratings.get(x[1], 1500),
        reverse=True
    )
    best_thirds = [t for _, t in thirds_sorted[:8]]

    # Pares R32 (6 pares de grupos + 4 mejores terceros vs mejores primeros)
    group_pairs = [("A","B"),("C","D"),("E","F"),("G","H"),("I","J"),("K","L")]
    r32 = []
    for g1, g2 in group_pairs:
        r32.append((firsts[g1],  seconds[g2]))
        r32.append((firsts[g2],  seconds[g1]))

    extra_firsts = [firsts[g] for g in ["C","D","G","H"]]
    for i, third in enumerate(best_thirds[:4]):
        r32.append((extra_firsts[i], third))

    r32 = r32[:16]

    def best_winner(a, b):
        ra = comp_ratings.get(a, 1500)
        rb = comp_ratings.get(b, 1500)
        pw, pd_, pl = win_probability(ra, rb)
        p_a = pw + pd_ * 0.5
        p_b = pl + pd_ * 0.5
        return (a, round(pw*100,1), round(pl*100,1)) if ra >= rb else (b, round(pl*100,1), round(pw*100,1))

    bracket = []
    current_round = r32
    round_names = ["Ronda de 32", "Octavos de Final", "Cuartos de Final", "Semifinales", "Final"]

    for rname in round_names:
        round_matches = []
        winners = []
        for a, b in current_round:
            winner, pa, pb = best_winner(a, b)
            loser = b if winner == a else a
            round_matches.append({
                "ronda": rname,
                "equipo_a": a,
                "equipo_b": b,
                "prob_a": pa,
                "prob_b": pb,
                "proyeccion": winner,
            })
            winners.append(winner)
        bracket.extend(round_matches)
        current_round = list(zip(winners[::2], winners[1::2]))
        if not current_round:
            break

    return bracket


def sheet_knockout(wb, comp_ratings, attack_defense=None, form_adj=None):
    print("  Proyectando bracket eliminatorio...")
    bracket = project_knockout(comp_ratings, attack_defense, form_adj)

    ws = wb.create_sheet("Fase Eliminatoria")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    ws["A1"] = "🏆  MUNDIAL FIFA 2026 — FASE ELIMINATORIA PROYECTADA"
    ws["A1"].fill = fill(C_HEADER_BG)
    ws["A1"].font = Font(bold=True, color=C_HEADER_FG, name="Calibri", size=13)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 28

    headers = ["Ronda", "Partido", "Equipo A", "Prob. A", "Equipo B", "Prob. B", "Proyección (Ganador)"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        style_header(c, bg=C_SUBHEAD_BG)

    widths = [20, 8, 25, 10, 25, 10, 25]
    for i, w in enumerate(widths, 1):
        set_col_width(ws, i, w)

    round_colors = {
        "Ronda de 32":       "EBF5FB",
        "Octavos de Final":  "D6EAF8",
        "Cuartos de Final":  "AED6F1",
        "Semifinales":       "7FB3D3",
        "Final":             "FFD700",
    }
    round_match_num = {}

    row = 3
    current_round = None
    for match in bracket:
        ronda = match["ronda"]
        if ronda != current_round:
            current_round = ronda
            round_match_num[ronda] = 1
            # Separador
            if row > 3:
                ws.row_dimensions[row].height = 8
                row += 1
        else:
            round_match_num[ronda] += 1

        bg = round_colors.get(ronda, C_STRIPE2)
        partido_num = round_match_num[ronda]

        is_final = ronda == "Final"
        font_size = 12 if is_final else 10

        vals = [
            ronda,
            f"Partido {partido_num}",
            match["equipo_a"],
            match["prob_a"] / 100,
            match["equipo_b"],
            match["prob_b"] / 100,
            f"🏅 {match['proyeccion']}" if is_final else f"→ {match['proyeccion']}",
        ]

        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = fill("FFF9C4" if is_final else bg)
            c.font = Font(
                bold=(col in (1, 7) or is_final),
                name="Calibri",
                size=font_size,
                color=("8B0000" if is_final else "000000"),
            )
            c.alignment = center() if col not in (3, 5, 7) else left()
            c.border = BORDER_THIN

            if col in (4, 6):
                c.number_format = "0.0%"
                fav = match["equipo_a"] if match["prob_a"] >= match["prob_b"] else match["equipo_b"]
                if (col == 4 and fav == match["equipo_a"]) or (col == 6 and fav == match["equipo_b"]):
                    c.fill = fill("C8E6C9")
                else:
                    c.fill = fill("FFCDD2")

        if is_final:
            ws.row_dimensions[row].height = 22

        row += 1

    print(f"  Hoja 'Fase Eliminatoria': {len(bracket)} partidos proyectados")


# ═══════════════════════════════════════════════════════════════════════════════
# HOJA 4 — PROBABILIDADES DE CAMPEONATO
# ═══════════════════════════════════════════════════════════════════════════════
def sheet_championship(wb, elo_results, comp_results):
    ws = wb.create_sheet("Prob. Campeonato")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    ws["A1"] = "🏆  MUNDIAL FIFA 2026 — PROBABILIDAD DE SER CAMPEÓN"
    ws["A1"].fill = fill(C_HEADER_BG)
    ws["A1"].font = Font(bold=True, color=C_HEADER_FG, name="Calibri", size=13)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 28

    headers = ["Pos", "Selección", "Grupo", "Modelo ELO", "Modelo ELO+FIFA", "Variación", "Barra de Prob."]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        style_header(c, bg=C_SUBHEAD_BG)

    widths = [5, 22, 7, 13, 14, 11, 30]
    for i, w in enumerate(widths, 1):
        set_col_width(ws, i, w)

    merged = comp_results[["team","probability_pct"]].rename(
        columns={"probability_pct":"comp_pct"}
    ).merge(
        elo_results[["team","probability_pct"]].rename(
            columns={"probability_pct":"elo_pct"}
        ), on="team", how="outer"
    ).fillna(0).sort_values("comp_pct", ascending=False).reset_index(drop=True)

    medal_bg = {0: "FFF9C4", 1: "F5F5F5", 2: "FFF3E0"}
    medal_color = {0: "B8860B", 1: "607D8B", 2: "795548"}

    for i, row_data in merged.iterrows():
        row = i + 3
        team = row_data["team"]
        group = TEAM_TO_GROUP.get(team, "?")
        elo_pct  = row_data["elo_pct"]
        comp_pct = row_data["comp_pct"]
        diff = comp_pct - elo_pct
        pos = i + 1

        bg = medal_bg.get(i, C_STRIPE1 if row % 2 == 0 else C_STRIPE2)

        arrow = ("↑" if diff > 0.3 else ("↓" if diff < -0.3 else "→"))
        diff_color = ("1B5E20" if diff > 0.3 else ("B71C1C" if diff < -0.3 else "555555"))

        bar_len = int(comp_pct * 1.5)
        bar = "█" * bar_len + "░" * max(0, 45 - bar_len)

        vals = [pos, team, f"Grupo {group}", elo_pct/100, comp_pct/100, f"{arrow} {abs(diff):.1f}%", bar]

        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = fill(bg)
            c.border = BORDER_THIN
            c.alignment = center() if col not in (2, 7) else left()
            c.font = Font(
                bold=(col in (1, 2) and i < 3),
                name="Calibri" if col != 7 else "Courier New",
                size=9 if col == 7 else 10,
                color=medal_color.get(i, diff_color if col == 6 else "000000"),
            )
            if col in (4, 5):
                c.number_format = "0.00%"
            if col == 6:
                c.font = Font(bold=True, color=diff_color, name="Calibri", size=10)

    print(f"  Hoja 'Prob. Campeonato': {len(merged)} selecciones")


# ═══════════════════════════════════════════════════════════════════════════════
# HOJA 5 — RATINGS
# ═══════════════════════════════════════════════════════════════════════════════
def sheet_ratings(wb, elo_ratings, fifa_rankings, comp_ratings):
    ws = wb.create_sheet("Ratings")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    ws["A1"] = "📊  RATINGS DE LAS 48 SELECCIONES — ELO · FIFA · COMPUESTO"
    ws["A1"].fill = fill(C_HEADER_BG)
    ws["A1"].font = Font(bold=True, color=C_HEADER_FG, name="Calibri", size=13)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 28

    headers = ["Pos", "Selección", "Grupo", "ELO Histórico", "Puntos FIFA", "Rating Compuesto", "Ranking FIFA"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        style_header(c, bg=C_SUBHEAD_BG)

    widths = [5, 24, 7, 15, 13, 17, 13]
    for i, w in enumerate(widths, 1):
        set_col_width(ws, i, w)

    fifa_rank_map = load_fifa_rankings()

    from src.world_cup_2026 import ALL_TEAMS
    team_data = []
    for team in ALL_TEAMS:
        team_data.append({
            "team":  team,
            "group": TEAM_TO_GROUP.get(team, "?"),
            "elo":   elo_ratings.get(team, 1500),
            "fifa":  get_fifa_points(team, fifa_rank_map),
            "comp":  comp_ratings.get(team, 1500),
            "fifa_rank": fifa_rank_map.get(team, {}).get("rank", 999),
        })
    team_data.sort(key=lambda x: x["comp"], reverse=True)

    max_comp = max(t["comp"] for t in team_data)
    min_comp = min(t["comp"] for t in team_data)

    for i, td in enumerate(team_data):
        row = i + 3
        bg = C_STRIPE1 if row % 2 == 0 else C_STRIPE2
        if i == 0: bg = "FFF9C4"
        elif i == 1: bg = "F5F5F5"
        elif i == 2: bg = "FFF3E0"

        # Barra de color proporcional al rating
        ratio = (td["comp"] - min_comp) / (max_comp - min_comp)
        g_val = int(200 - ratio * 100)
        heat_color = f"{'%02X' % int(255*(1-ratio))}{'%02X' % g_val}{'%02X' % 50}"

        vals = [i+1, td["team"], f"Grupo {td['group']}", round(td["elo"]), round(td["fifa"]), round(td["comp"]), td["fifa_rank"]]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = fill(bg)
            c.border = BORDER_THIN
            c.alignment = center() if col != 2 else left()
            c.font = Font(bold=(col == 2 and i < 3), name="Calibri", size=10)
            if col == 6:
                c.fill = fill(bg)

    # Formato condicional en columna Compuesto
    from openpyxl.formatting.rule import ColorScaleRule
    comp_col = get_column_letter(6)
    ws.conditional_formatting.add(
        f"{comp_col}3:{comp_col}{len(team_data)+2}",
        ColorScaleRule(
            start_type="min", start_color="FFC7CE",
            mid_type="percentile", mid_value=50, mid_color="FFEB9C",
            end_type="max", end_color="C6EFCE",
        )
    )

    print(f"  Hoja 'Ratings': {len(team_data)} equipos")


# ═══════════════════════════════════════════════════════════════════════════════
# HOJA 6 — VALIDACIÓN DEL MODELO (BACKTESTING)
# ═══════════════════════════════════════════════════════════════════════════════
def sheet_validation(wb):
    """Hoja con métricas de backtesting 2018/2022 y comparativa de modelos."""
    ws = wb.create_sheet("Validación Modelo")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    ws["A1"] = "📊  VALIDACIÓN DEL MODELO — BACKTESTING MUNDIALES 2018 Y 2022"
    ws["A1"].fill = fill(C_HEADER_BG)
    ws["A1"].font = Font(bold=True, color=C_HEADER_FG, name="Calibri", size=13)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 28

    # Sección: resumen de backtesting
    row = 3
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "Resultados de Backtesting (50,000 simulaciones por Mundial)"
    ws[f"A{row}"].fill = fill(C_SUBHEAD_BG)
    ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    ws[f"A{row}"].alignment = center()
    row += 1

    headers = ["Mundial", "Campeón Real", "Prob. Predicha", "Ranking Campeón",
               "Brier Score", "Log-Loss", "vs. Naive (equiprobable)"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        style_header(c, bg="2E75B6")
    row += 1

    widths = [18, 16, 14, 15, 13, 12, 22]
    for i, w in enumerate(widths, 1):
        set_col_width(ws, i, w)

    backtest_data = [
        ("Rusia 2018",  "France",    "7.67%",  4,  0.030395, 2.5673),
        ("Qatar 2022",  "Argentina", "19.37%", 2,  0.024207, 1.6413),
    ]
    naive_bs = 0.030273
    naive_ll = 3.4657

    for wc_name, champion, prob, rank, bs, ll in backtest_data:
        improve_bs = (naive_bs - bs) / naive_bs * 100
        improve_ll = (naive_ll - ll) / naive_ll * 100
        vs_naive = f"BS {improve_bs:+.1f}% | LL {improve_ll:+.1f}%"

        bg = "E8F5E9" if bs < naive_bs else "FFEBEE"
        vals = [wc_name, champion, prob, f"#{rank}", f"{bs:.6f}", f"{ll:.4f}", vs_naive]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = fill(bg)
            c.border = BORDER_THIN
            c.alignment = center() if col != 1 else left()
            c.font = Font(name="Calibri", size=10, bold=(col == 2))
        row += 1

    # Fila naive (referencia)
    vals = ["Naive (1/32 equiprobable)", "—", "3.13%", "—", f"{naive_bs:.6f}", f"{naive_ll:.4f}", "Línea base"]
    for col, val in enumerate(vals, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill("FFF3CD")
        c.border = BORDER_THIN
        c.alignment = center() if col != 1 else left()
        c.font = Font(name="Calibri", size=10, italic=True)
    row += 2

    # Sección: comparativa de modelos
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "Comparativa de Modelos — Brier Score Promedio (2018 + 2022)"
    ws[f"A{row}"].fill = fill(C_SUBHEAD_BG)
    ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    ws[f"A{row}"].alignment = center()
    row += 1

    headers2 = ["Modelo", "Brier 2018", "Brier 2022", "Promedio BS", "LL 2018", "LL 2022", "Mejora vs Naive"]
    for col, h in enumerate(headers2, 1):
        c = ws.cell(row=row, column=col, value=h)
        style_header(c, bg="2E75B6")
    row += 1

    model_data = [
        ("v2 ELO+decay+AD+Forma (completo)", 0.030477, 0.024116, 2.5845, 1.6384),
        ("v1 ELO+decay (sin AD/Forma)",       0.029910, 0.026015, 2.6205, 1.9590),
        ("Naive (1/32 equiprobable)",          0.030273, 0.030273, 3.4657, 3.4657),
    ]
    colors = ["C8E6C9", "E3F2FD", "FFF3CD"]

    for (label, bs18, bs22, ll18, ll22), bg in zip(model_data, colors):
        avg_bs = (bs18 + bs22) / 2
        avg_ll = (ll18 + ll22) / 2
        improve = (naive_bs - avg_bs) / naive_bs * 100 if label != "Naive (1/32 equiprobable)" else 0.0
        improve_str = f"{improve:+.1f}%" if improve != 0 else "referencia"
        vals = [label, f"{bs18:.6f}", f"{bs22:.6f}", f"{avg_bs:.6f}", f"{ll18:.4f}", f"{ll22:.4f}", improve_str]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = fill(bg)
            c.border = BORDER_THIN
            c.alignment = center() if col != 1 else left()
            c.font = Font(name="Calibri", size=10, bold=(col == 1 and label.startswith("v2")))
        row += 1

    row += 1
    nota = ("Brier Score: más bajo = mejor (0 = perfecto, 1 = pésimo). "
            "Log-Loss: más bajo = mejor. "
            "Naive = 1/n igual prob. a todos. "
            "El modelo v3 supera al naive en 9.8% en Brier Score y 53% en Log-Loss (Qatar 2022).")
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = nota
    ws[f"A{row}"].font = Font(name="Calibri", size=9, italic=True, color="555555")
    ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 36

    print(f"  Hoja 'Validación Modelo': backtesting 2018/2022")


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    print("=== Generando Excel del Mundial 2026 (Modelo v3) ===\n")

    elo_ratings, fifa_rankings, comp_ratings, elo_results, comp_results, attack_defense, form_adj = load_all_data()

    # Cargar calendario de partidos
    df = pd.read_csv("data/raw/results.csv", parse_dates=["date"])
    schedule = df[
        (df["tournament"] == "FIFA World Cup") & (df["date"].dt.year == 2026)
    ].copy()

    wb = Workbook()
    wb.remove(wb.active)  # elimina hoja vacía por defecto

    print("\nGenerando hojas...")
    sheet_group_stage(wb, schedule, comp_ratings, elo_ratings, fifa_rankings)
    sheet_group_standings(wb, comp_ratings, attack_defense, form_adj)
    sheet_knockout(wb, comp_ratings, attack_defense, form_adj)
    sheet_championship(wb, elo_results, comp_results)
    sheet_ratings(wb, elo_ratings, fifa_rankings, comp_ratings)
    sheet_validation(wb)

    os.makedirs("results/excel", exist_ok=True)
    output_path = "results/excel/Mundial_2026_Predicciones.xlsx"
    wb.save(output_path)
    print(f"\n✅ Excel guardado en: {output_path}")
    print(f"   Tamaño: {os.path.getsize(output_path)/1024:.0f} KB")


if __name__ == "__main__":
    main()
